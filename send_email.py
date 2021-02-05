#处理Excel的方法
import smtplib
from email.mime.text import MIMEText
from email.header import Header
from email import encoders
from email.mime.base import MIMEBase
from email.header import Header
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.utils import parseaddr, formataddr
import datetime
from openpyxl import load_workbook
import os

import logging
import logging.config
# CON_LOG='config\\' \
#         '.conf'
CON_LOG='config\\log.conf'
logging.config.fileConfig(CON_LOG)
logging=logging.getLogger()

def my_send_email(msg_to,device_model,device_version,number_all,wake_num,identify_number,file_path):
    '''
    msg_to:收件人邮箱列表
    device_model:设备类型
    device_version:版本号
    number_all:总运行次数
    wake_num:唤醒次数
    identify_number:识别次数
    file_path:excel结果文件路径
    '''
    now_time = datetime.datetime.now()
    year = now_time.year
    month = now_time.month
    day = now_time.day
    mytime = str(year) + " 年 " + str(month) + " 月 " + str(day) + " 日 "
    msg_from = '1508691067@qq.com'  # 发送方邮箱
    passwd = 'fgaplzfksqsihdbe'

    # msg_from = '319910390@qq.com'  # 发送方邮箱
    # passwd='xmptiwfdopslbigj'
    #msg_to = ['1508691067@qq.com', '319910390@qq.com']

    subject = '语音识别测试结果'
    #contentall = ''

    wake_proportion = format(int(wake_num) / int(number_all), '.1%')  #唤醒率
    identify_proportion = format(int(identify_number) / int(number_all), '.1%')  #识别率

    #构造要发送的内容格式
    content = '''
                <html>
                <body>
                    <h1 align="center">智能音响语音识别测试结果</h1>
                    <p><strong>您好：</strong></p>
                    <blockquote><p><strong>以下数据是语音识别测试结果,请查收！</strong></p></blockquote>

                    <blockquote><p><strong>型号：{device_model}&nbsp &nbsp &nbsp &nbsp &nbsp &nbsp 版本:{device_version} </strong></p></blockquote>
                    <blockquote><p><strong>运行次数：{number_all} &nbsp &nbsp &nbsp &nbsp &nbsp &nbsp &nbsp 唤醒次数:{wake_num} &nbsp &nbsp &nbsp &nbsp &nbsp 识别正确次数{identify_number}</strong></p></blockquote>
                    <blockquote><p><strong>唤醒率: {wake_proportion} &nbsp &nbsp &nbsp &nbsp &nbsp  识别率:{identify_proportion}</strong></p></blockquote>
                    <p align="right">{mytime}</p>
                <body>
                <html>
                '''.format(device_model=device_model, device_version=device_version, number_all=number_all,
                           wake_num=wake_num, identify_number=identify_number, wake_proportion=wake_proportion,
                           identify_proportion=identify_proportion, mytime=mytime)
    # 这个应该就是构建了要给html对象
    msg = MIMEMultipart()
    msg.attach(MIMEText(content, 'html', 'utf-8'))

    #加入excel附件
    # part=MIMEBase('application','octet-stream')
    # part.set_payload(open("test.xls","rb").read())
    # encoders.encode_base64(part)
    # part.add_header('Content-Disposition','attacment;fileneme="test.xls"')
    # msg.attach(part)

    att1=MIMEText(open(file_path,'rb').read(),'base64','utf-8')
    att1["Content-Type"]='application/octet-stream'
    #att1['Content-Disposition']='attachment;filename='+file_path
    excel_base_path=os.path.dirname(file_path)  #获取的是路径
    excel_base_name=os.path.basename(file_path)  #获取到的是文件名称
    att1['Content-Disposition'] = 'attachment;filename=' + excel_base_name
    msg.attach(att1)

    # 放入邮件主题
    msg['Subject'] = subject

    # 放入发件人,这是展示在邮件里面的，和时间的发件人没有关系
    msg['From'] = msg_from
    try:
        # 通过ssl方式发送，服务器地址，端口
        s = smtplib.SMTP_SSL("smtp.qq.com", 465)
        #登录邮箱
        s.login(msg_from, passwd)
        s.sendmail(msg_from, msg_to, msg.as_string())
        #print("邮件发送成功")
        logging.info("邮件发送成功")
    except s.SMTPException as e:
        #print(e)
        logging.error(e)
    finally:
        s.quit()


def my_send(excel_path,msg_to,device_version=None):
    myexcel=load_workbook(excel_path)
    #print(myexcel)
    mysheet=myexcel.active
    #print(mysheet)
    max_row=mysheet.max_row  #行数
    #print("max_column:",max_row)

    #有效行数，除去表头剩余的行数
    row_num=max_row-2
    # print('row_num:',row_num)
    # res=mysheet.cell(1,1).value
    # print(res)
    #唤醒次数
    wake_nmu=0

    #识别次数
    identify_number=0
    for i in range(row_num):
        is_wake=mysheet.cell(i+3,7).value
        is_identify=mysheet.cell(i+3,9).value

        print("is_wake: ",is_wake)
        print("type(is_wake): ",type(is_wake))

        print("is_identify: ",is_identify)
        print("type(is_identify): ",type(is_identify))

        if is_wake:
            wake_nmu=wake_nmu+1
        if is_identify:
            identify_number=identify_number+1

    #print("wake_num:",wake_nmu)
    #print("identify_number:",identify_number)
    logging.info("wake_num: "+str(wake_nmu))
    logging.info("identify_number:"+str(identify_number))

    #发送邮件参数
    msg_to=msg_to
    device_model = '智能音响'
    if device_version==None:
        device_version = ' '
    number_all = row_num
    wake_num =wake_nmu
    identify_number=identify_number
    my_send_email(msg_to, device_model, device_version, number_all, wake_num, identify_number, excel_path)



if __name__=="__main__":

    excel_path="result_2020_04_15_14_19_39.xlsx"
    msg_to = ['1508691067@qq.com']
    my_send(excel_path,msg_to)

    # '''
    # msg_to:收件人邮箱列表
    # device_model:设备类型
    # device_version:版本号
    # number_all:总运行次数
    # wake_num:唤醒次数
    # identify_number:识别次数
    # '''
    # msg_to=['1508691067@qq.com']
    # device_model='智能音响'
    # device_version='1.1.1'
    # number_all=100
    # wake_num=90
    # identify_number=80
    # my_send_email(msg_to,device_model,device_version,number_all,wake_num,identify_number,'test.xls')





