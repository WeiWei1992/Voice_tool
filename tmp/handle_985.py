from openpyxl import load_workbook

jiubawu_list=[]
eryiyi_list=[]
with open('211.txt',encoding='utf-8') as f:
    lines=f.readlines()
    for line in lines:
        line=line.strip()
        eryiyi_list.append(line)
with open('985.txt',encoding='utf-8') as f:
    lines=f.readlines()
    for line in lines:
        line = line.strip()
        jiubawu_list.append(line)

print(jiubawu_list)
print(eryiyi_list)

#95所一流学科建设高校名单
def handle_excel(file,sheetname,column,col):
    wb=load_workbook(file)
    sheet=wb[sheetname]

    max_column=sheet.max_row

    for i in range(max_column):
        schllo_name=sheet.cell(i+1,column).value
        if schllo_name in eryiyi_list:
            sheet.cell(i+1,col).value='211'


    for i in range(max_column):
        schllo_name=sheet.cell(i+1,column).value
        if schllo_name in jiubawu_list:
            sheet.cell(i+1,col).value='985/211'


    print(max_column)
    print(sheet.cell(4,2).value)
    wb.save("排名.xlsx")

if __name__=="__main__":
    file="排名.xlsx"
    sheetname=""
    column=2
    col=5
    handle_excel(file,sheetname,column,col)