import openpyxl
from openpyxl import load_workbook
from datetime import datetime
import os
import time
from openpyxl.styles import PatternFill
from openpyxl.styles import Color, Font, Alignment

import logging
import logging.config
CON_LOG='config\\log.conf'
logging.config.fileConfig(CON_LOG)
logging=logging.getLogger()


def creat_excel(filename=None):
    dt = datetime.now()
    now_time = dt.strftime('%Y_%m_%d_%H_%M_%S')  # 得用下划线，用： 号无法截图保存
    my_path = os.path.abspath(os.getcwd())
    if filename==None:
        filename = my_path + '/Result/result_%s.xlsx' % (now_time)
    else:#为了gui加上的，ui页面中要选中路径，传进来的是路径，不包括文件名
        filename=filename+'/result_%s.xlsx' %(now_time)

    logging.info("excel路径"+str(filename))
    wb=openpyxl.Workbook()
    mysheet=wb.active

    mysheet.merge_cells('A1:L1')
    mysheet.cell(row=1,column=1,value="智能音箱语音自动化测试结果")
    mysheet.row_dimensions[1].height = 25

    # 然后如下设置：
    # 设置表头字体居中
    mycell=mysheet['A1']
    mycell.font=Font(name=u'宋体', bold=True)
    mycell.alignment=Alignment(horizontal='center', vertical='center')

    result_head=['唤醒词路径','交互词路径','交互文本路径','全日志路径','截取后的日志路径','audio文件路径','唤醒结果','唤醒词语句','交互结果','交互语音识别文本','实际的语音文本','NLP返回结果']
    for i,item in enumerate(result_head):
        #print(i,item)
        mysheet.cell(row=2, column=i + 1, value=item).alignment=Alignment(horizontal='center',vertical='center')
    mysheet['G2'].font=Font(name=u'宋体',bold=True)
    mysheet['I2'].font=Font(name=u'宋体',bold=True)

    mysheet.column_dimensions['A'].width = 50
    mysheet.column_dimensions['B'].width = 50
    mysheet.column_dimensions['C'].width = 50
    mysheet.column_dimensions['D'].width = 50
    mysheet.column_dimensions['E'].width = 50
    mysheet.column_dimensions['F'].width = 50
    mysheet.column_dimensions['G'].width = 20
    mysheet.column_dimensions['H'].width = 20
    mysheet.column_dimensions['K'].width = 20
    mysheet.column_dimensions['I'].width = 50
    mysheet.column_dimensions['J'].width = 50
    mysheet.column_dimensions['K'].width = 50
    mysheet.column_dimensions['L'].width = 100
    mysheet.row_dimensions[2].height=25

    mysheet.title="测试结果"
    #mysheet.row_dimensions[3].height=25  #设置行高,设置第3行的行高

    wb.save(filename)
    print(filename)
    return filename

def write_excel(filename,wake_path,inter_path,text_path,all_log_path,intercept_log_path,audio_path,wake_result,wake_txt,inter_result,inter_idently_text,real_text,nlp_text):
    wb=load_workbook(filename)
    sheet=wb.active
    i=sheet.max_row
    #column_num=sheet.max_column
    # print(sheet.max_row)
    # print(sheet.max_column)
    sheet.row_dimensions[i+1].height = 25
    #green_fill = PatternFill(fill_type="solid", fgColor="AACF91")

    #sheet.row_dimensions[1].fill = green_fill
    # print(wake_result)
    # print("type(wake_result):",type(wake_result))
    if wake_result==False or inter_result==False:
        #sheet.row_dimensions[i+1].fill=green_fill
        #填充颜色
        red_fill = PatternFill(fill_type='solid', fgColor="ff441f")
        sheet.cell(row=i + 1, column=1, value=wake_path).fill=red_fill
        sheet.cell(row=i + 1, column=2, value=inter_path).fill=red_fill
        sheet.cell(row=i + 1, column=3, value=text_path).fill=red_fill
        sheet.cell(row=i + 1, column=4, value=all_log_path).fill=red_fill
        sheet.cell(row=i + 1, column=5, value=intercept_log_path).fill=red_fill
        sheet.cell(row=i+1,column=6,value=audio_path).fill=red_fill
        sheet.cell(row=i + 1, column=7, value=wake_result).fill=red_fill
        sheet.cell(row=i + 1, column=8, value=wake_txt).fill=red_fill
        sheet.cell(row=i + 1, column=9, value=inter_result).fill=red_fill
        sheet.cell(row=i + 1, column=10, value=inter_idently_text).fill=red_fill
        sheet.cell(row=i + 1, column=11,value=real_text).fill=red_fill
        sheet.cell(row=i + 1, column=12, value=nlp_text).fill=red_fill
    else:
        green_file=PatternFill(fill_type='solid',fgColor="16feb5")
        sheet.cell(row=i + 1, column=1, value=wake_path).fill=green_file
        sheet.cell(row=i + 1, column=2, value=inter_path).fill=green_file
        sheet.cell(row=i + 1, column=3, value=text_path).fill=green_file
        sheet.cell(row=i + 1, column=4, value=all_log_path).fill=green_file
        sheet.cell(row=i + 1, column=5, value=intercept_log_path).fill=green_file
        sheet.cell(row=i+1,column=6,value=audio_path).fill=green_file
        sheet.cell(row=i + 1, column=7, value=wake_result).fill=green_file
        sheet.cell(row=i + 1, column=8, value=wake_txt).fill=green_file
        sheet.cell(row=i + 1, column=9, value=inter_result).fill=green_file
        sheet.cell(row=i + 1, column=10, value=inter_idently_text).fill=green_file
        sheet.cell(row=i + 1, column=11,value=real_text).fill=green_file
        sheet.cell(row=i + 1, column=12, value=nlp_text).fill=green_file

    #设置居中,结果加重显示
    sheet.cell(row=i + 1, column=1).alignment = Alignment(horizontal='center', vertical='center')
    sheet.cell(row=i + 1, column=2).alignment = Alignment(horizontal='center', vertical='center')
    sheet.cell(row=i + 1, column=3).alignment = Alignment(horizontal='center', vertical='center')
    sheet.cell(row=i + 1, column=4).alignment = Alignment(horizontal='center', vertical='center')
    sheet.cell(row=i + 1, column=5).alignment = Alignment(horizontal='center',
                                                                                    vertical='center')
    sheet.cell(row=i + 1, column=6).alignment = Alignment(horizontal='center', vertical='center')
    #sheet.cell(row=i + 1, column=6).font = Font(name=u'宋体', bold=True)
    sheet.cell(row=i + 1, column=7).alignment = Alignment(horizontal='center', vertical='center')
    sheet.cell(row=i + 1, column=7).font = Font(name=u'宋体', bold=True)
    sheet.cell(row=i + 1, column=8).alignment = Alignment(horizontal='center', vertical='center')
    #sheet.cell(row=i + 1, column=8).font = Font(name=u'宋体', bold=True)

    sheet.cell(row=i + 1, column=9).alignment = Alignment(horizontal='center', vertical='center')
    sheet.cell(row=i + 1, column=9).font = Font(name=u'宋体', bold=True)
    sheet.cell(row=i + 1, column=10).alignment = Alignment(horizontal='center', vertical='center')
    sheet.cell(row=i + 1, column=11).alignment = Alignment(horizontal='center', vertical='center')
    sheet.cell(row=i + 1, column=12).alignment = Alignment(horizontal='center', vertical='center')

    wb.save(filename)


if __name__=="__main__":
    filename=creat_excel()
    #filename='C:\\Users\\weiwei\\PycharmProjects\\test_appium\\Audio_Accuracy\\Result\\result_2020_04_13_12_38_54.xlsx'
    write_excel(filename,'my_wake_path','my_inter_path','my_text_path','my_all_log_path','my_intercept_log_path',False,'my_wake_txt',False,'my_inter_identify_text','my_real_text','my_nlp_text')

    write_excel(filename, 'my_wake_path', 'my_inter_path', 'my_text_path', 'my_all_log_path', 'my_intercept_log_path',
                True, 'my_wake_txt', True, 'my_inter_text','my_real_text', 'my_nlp_text')
    write_excel(filename, 'my_wake_path', 'my_inter_path', 'my_text_path', 'my_all_log_path', 'my_intercept_log_path',
                False, 'my_wake_txt', True, 'my_inter_text','my_real_text', 'my_nlp_text')
    write_excel(filename, 'my_wake_path', 'my_inter_path', 'my_text_path', 'my_all_log_path', 'my_intercept_log_path',
                True, 'my_wake_txt', False, 'my_inter_text','my_real_text', 'my_nlp_text')
    write_excel(filename, 'my_wake_path', 'my_inter_path', 'my_text_path', 'my_all_log_path', 'my_intercept_log_path',
                True, 'my_wake_txt', True, 'my_inter_text','my_real_text', 'my_nlp_text')
    #write_excel(filename, 'my_wake_path', 'my_inter_path', 'my_text_path', 'my_all_log_path', 'my_intercept_log_path',
